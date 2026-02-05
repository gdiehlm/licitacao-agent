import shutil
from dataclasses import dataclass
from typing import Dict, List
from datetime import datetime
from zoneinfo import ZoneInfo
import openpyxl

@dataclass
class OutputItem:
    prioridade: int
    descricao_resumida: str
    descricao_detalhada: str
    unidade: str
    quantidade: int
    preco_unit: float

def col_to_index(col: str) -> int:
    col = col.upper()
    n = 0
    for ch in col:
        n = n * 26 + (ord(ch) - ord('A') + 1)
    return n

def clear_rows(ws, start_row: int, columns: Dict[str,str], max_clear: int = 300):
    cols = [columns[k] for k in ["prioridade","descricao_resumida","descricao_detalhada","unidade","quantidade","preco_unit","preco_total"]]
    for r in range(start_row, start_row + max_clear):
        if r > start_row and all(ws.cell(r, col_to_index(c)).value in (None, "") for c in cols):
            break
        for c in cols:
            ws.cell(r, col_to_index(c)).value = None

def write_from_template(template_path: str, out_path: str, sheet: str, start_row: int, columns: Dict[str,str], items: List[OutputItem]) -> str:
    shutil.copy(template_path, out_path)
    wb = openpyxl.load_workbook(out_path)
    ws = wb[sheet]

    clear_rows(ws, start_row, columns)

    for i, it in enumerate(items):
        r = start_row + i
        ws.cell(r, col_to_index(columns["prioridade"])).value = it.prioridade
        ws.cell(r, col_to_index(columns["descricao_resumida"])).value = it.descricao_resumida
        ws.cell(r, col_to_index(columns["descricao_detalhada"])).value = it.descricao_detalhada
        ws.cell(r, col_to_index(columns["unidade"])).value = it.unidade
        ws.cell(r, col_to_index(columns["quantidade"])).value = it.quantidade
        ws.cell(r, col_to_index(columns["preco_unit"])).value = float(it.preco_unit)
        ws.cell(r, col_to_index(columns["preco_total"])).value = f"={columns['preco_unit']}{r}*{columns['quantidade']}{r}"
        ws.cell(r, col_to_index(columns["preco_unit"])).number_format = '"R$" #,##0.00'
        ws.cell(r, col_to_index(columns["preco_total"])).number_format = '"R$" #,##0.00'

    wb.save(out_path)
    return out_path

def dated_filename(prefix: str = "Pregão") -> str:
    tz = ZoneInfo("America/Sao_Paulo")
    stamp = datetime.now(tz).strftime("%d-%m-%y")
    return f"{prefix} {stamp}.xlsx"
